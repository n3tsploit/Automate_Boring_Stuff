import sys
import openpyxl
from openpyxl.styles import Font

number = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active
font_obj = Font(bold=True)
for i in range(2, number+2):
    sheet.cell(row=1, column=i).value = i - 1
    sheet.cell(row=1, column=i).font = font_obj
    sheet.cell(row=i, column=1).value = i - 1
    sheet.cell(row=i, column=1).font = font_obj

for row in range(2, number+2):
    for column in range(2, number+2):
        sheet.cell(row=row, column=column).value = sheet.cell(row=1, column=column).value * sheet.cell(row=row,
                                                                                                       column=1).value
wb.save('test.xlsx')

"""
Create a program multiplicationTable.py that takes a number N from the command line and creates an
multiplication table in an Excel spreadsheet. For example, when the program is run like this:
N × N
py multiplicationTable.py 6

"""