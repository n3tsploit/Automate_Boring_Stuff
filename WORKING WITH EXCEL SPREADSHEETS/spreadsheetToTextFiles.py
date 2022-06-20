import openpyxl

wb = openpyxl.load_workbook('ls.xlsx')
sheet = wb.active

for column in range(1, sheet.max_column + 1):
    with open(f'output_{column}.txt', 'w') as w:
        for row in range(1, sheet.max_row+1):
            values = sheet.cell(row=row, column=column).value
            if values is not None:
                w.write(str(values))
                w.write('\n')

    print(f'Done with file {column}')

"""
Write a program that performs the tasks of the previous program in reverse order: the program should open
a spreadsheet and write the cells of column A into one text file, the cells of column B into another text file,
and so on.

"""