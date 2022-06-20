import openpyxl

wb = openpyxl.load_workbook('test.xlsx')
new_wb = openpyxl.Workbook()
sheet = wb.active
new_sheet = new_wb.active
main_list = []
for i in range(1, sheet.max_row + 1):
    row_list = []
    for column in range(1, sheet.max_column + 1):
        values = sheet.cell(row=i, column=column).value
        row_list.append(values)
    main_list.append(row_list)

for row in range(1,len(main_list)+1):
    for column in range(1, len(main_list[row-1])+1):
        values = main_list[row-1][column-1]
        new_sheet.cell(row=column, column=row).value = values

new_wb.save('reverse1.xlsx')
print('Done')

"""
Write a program to invert the row and column of the cells in the spreadsheet. For example, the value at row
5, column 3 will be at row 3, column 5 (and vice versa). This should be done for all cells in the spreadsheet.

"""