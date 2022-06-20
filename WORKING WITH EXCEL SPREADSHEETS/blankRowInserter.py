import sys

import openpyxl

if len(sys.argv) == 4:
    row_number = int(sys.argv[1])
    blanks_number = int(sys.argv[2])
    file_name = sys.argv[3]

    wb = openpyxl.load_workbook(f'{file_name}')
    sheet = wb.active
    mb = openpyxl.Workbook()
    sheet1 = mb.active

    for row in range(1, sheet.max_row + 1):
        if row < row_number:
            for column in range(1, sheet.max_column + 1):
                values = sheet.cell(row=row, column=column).value
                sheet1.cell(row=row, column=column).value = values
        else:
            new_row = row + blanks_number
            for column in range(1, sheet.max_column + 1):
                values = sheet.cell(row=row, column=column).value
                sheet1.cell(row=new_row, column=column).value = values

    mb.save('new_' + file_name)

else:
    print('Input number of rows,number of balnks,filename')

"""
Create a program blankRowInserter.py that takes two integers and a filename string as command line
arguments. Let’s call the first integer N and the second integer M . Starting at row N , the program should
insert M blank rows into the spreadsheet. For example, when the program is run like this:
python blankRowInserter.py 3 2 myProduce.xlsx
"""
