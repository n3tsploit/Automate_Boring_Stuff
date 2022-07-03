import openpyxl
import csv
import os
from path import Path

os.makedirs('csv_files', exist_ok=True)
p = Path('.')
for file in list(p.glob('*.xlsx')):
    print(f'Extracting for {file}')
    wb = openpyxl.load_workbook(file)
    for sheet in wb.sheetnames:
        csv_content = open(f'./csv_files/{file.strip(".xlsx")}_{sheet}.csv', 'w', newline='')
        sheet_content = wb[sheet]
        for j in range(sheet_content.max_row):
            row = []
            for i in range(sheet_content.max_column):
                row.append(sheet_content.cell(row=j+1, column=i+1).value)
            csv_writer = csv.writer(csv_content)
            csv_writer.writerow(row)
        csv_content.close()
    print('Done!')

"""
Excel can save a spreadsheet to a CSV file with a few mouse clicks, but if
you had to convert hundreds of Excel files to CSVs, it would take hours
of clicking. Using the openpyxl module from Chapter 12, write a program
that reads all the Excel files in the current working directory and outputs
them as CSV files.
A single Excel file might contain multiple sheets; you’ll have to create
one CSV file per sheet . The filenames of the CSV files should be <excel
filename>_<sheet title>.csv , where <excel filename> is the filename of the
Excel file without the file extension (for example, 'spam_data' , not
'spam_data.xlsx' ) and <sheet title> is the string from the Worksheet object’s
title variable.
"""
